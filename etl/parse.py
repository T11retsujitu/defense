"""公表ファイルのパース(原文保持)。

様式(公共調達の適正化 平成18年財計第2017号 に基づく公表様式):
  行0: 表題 / 行1: 列見出し / 行2: 副見出し / 行3以降: データ
  列見出しはファイル間でほぼ共通だが、列順は見出し名で解決する(年度による列ずれ対応)。

対応フォーマット:
  parse_xlsx — 中央調達(防衛装備庁)のxlsx
  parse_pdf  — 地方調達(地方防衛局等)のPDF(電子生成・罫線付き表。付紙様式第3/第4)
    xlsx版との見出し文言差(「契約締結日」→「契約を締結した日」等)は
    見出しの空白除去+部分一致で吸収する。
"""
from __future__ import annotations

import datetime
import json
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl

# 見出し(空白除去後の部分一致) -> 内部フィールド名
HEADER_MAP = [
    ("物品役務等の名称", "title"),
    ("数量", "quantity"),
    ("単位", "unit"),
    ("契約担当官", "agency"),
    ("契約締結日", "contract_date"),
    ("契約を締結した日", "contract_date"),
    ("契約相手方", "company"),
    ("相手方の商号", "company"),
    ("法人番号", "corporate_number"),
    ("予定価格", "planned_price"),
    ("契約金額", "amount"),
    ("落札率", "award_rate"),
    ("一般競争入札・指名競争入札の別", "method_detail"),
    ("随意契約によること", "method_detail"),
    ("備考", "remarks"),
]


def _squash(text: str) -> str:
    """見出し照合用: 改行・空白を除去(PDFはセル内改行が多い)。"""
    return re.sub(r"\s+", "", text)


def _map_header(cells) -> dict[int, str]:
    """見出し行のセル列 -> 内部フィールド名の対応表。"""
    col_map: dict[int, str] = {}
    for ci, cell in enumerate(cells):
        if cell is None:
            continue
        text = _squash(str(cell))
        for needle, field in HEADER_MAP:
            if _squash(needle) in text:
                col_map[ci] = field
                break
    return col_map


@dataclass
class RawRow:
    row_index: int
    values: dict          # field -> 原文値(str化前)
    row_json: str         # 行全体(列順のまま)のJSON


class ParseError(Exception):
    pass


def _cell_to_jsonable(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return v


def parse_xlsx(path: Path) -> list[RawRow]:
    """1ファイルをRawRow列へ。ヘッダ行が見つからない場合はParseError。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    header_idx = None
    for i, row in enumerate(rows[:10]):
        if row and any(c and "物品役務" in str(c) for c in row):
            header_idx = i
            break
    if header_idx is None:
        raise ParseError(f"header row not found: {path}")

    col_map = _map_header(rows[header_idx])

    required = {"title", "contract_date", "company", "amount"}
    if not required.issubset(set(col_map.values())):
        raise ParseError(f"missing columns {required - set(col_map.values())}: {path}")

    out: list[RawRow] = []
    # ヘッダ直後の副見出し行(公益法人の場合の内訳)をスキップするため、
    # データ行は「件名列が非空」で判定する
    for ri in range(header_idx + 1, len(rows)):
        row = rows[ri]
        if not row:
            continue
        values = {}
        for ci, field in col_map.items():
            if ci < len(row):
                values.setdefault(field, row[ci])
        title = values.get("title")
        if title is None or str(title).strip() == "":
            continue
        if "物品役務" in str(title):  # 重複ヘッダ
            continue
        row_json = json.dumps([_cell_to_jsonable(c) for c in row], ensure_ascii=False)
        out.append(RawRow(row_index=ri, values=values, row_json=row_json))
    return out


def parse_pdf(path: Path) -> list[RawRow]:
    """地方調達の公表PDF(付紙様式第3=競争入札/第4=随意契約)をRawRow列へ。

    - 各ページに見出し付きの表が繰り返されるため、ページごとに列対応を解決する。
    - 副見出し行(公益法人の内訳)・件名が空の行はスキップする。
    - row_index はファイル全体での通し番号(取込の差分同期キー)。
    """
    import pdfplumber  # PDFソースを使わない構成でも他パーサが動くよう遅延import

    out: list[RawRow] = []
    row_index = 0
    found_header = False
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                col_map: dict[int, str] = {}
                for row in table:
                    cells = ["" if c is None else str(c) for c in row]
                    if not any(c.strip() for c in cells):
                        continue
                    if any("物品役務等の名称" in _squash(c) for c in cells):
                        col_map = _map_header(cells)
                        found_header = True
                        continue
                    if not col_map:
                        continue  # 表頭より前の表題行など
                    title = cells[0].strip()
                    if not title or "公益法人の区" in _squash("".join(cells)):
                        continue
                    values = {}
                    for ci, field in col_map.items():
                        if ci < len(row):
                            values.setdefault(field, row[ci])
                    row_json = json.dumps(cells, ensure_ascii=False)
                    out.append(RawRow(row_index=row_index, values=values, row_json=row_json))
                    row_index += 1
    if not found_header:
        raise ParseError(f"header row not found: {path}")
    required = {"title", "contract_date", "company", "amount"}
    if out:
        present = set().union(*(r.values.keys() for r in out))
        if not required.issubset(present):
            raise ParseError(f"missing columns {required - present}: {path}")
    return out


def parse_source_file(path: Path, file_format: str) -> list[RawRow]:
    """フォーマット別パーサの振り分け。"""
    if file_format == "pdf":
        return parse_pdf(path)
    return parse_xlsx(path)
